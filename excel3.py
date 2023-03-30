import openpyxl as excel

# 기존 엑셀 불러오기
book = excel.load_workbook("py_excel01.xlsx")
#엑셀 시트 첫번째 선택
sheet = book.worksheets[0]

sheet.cell(row=2, column=1, value="행열을 이용한 파이썬 엑셀 사례")

row_cell = sheet.cell(row=3, column=1)
row_cell.value = "행열을 이용한 파이썬 두번째 사례"

book.save("py_excel02.xlsx")